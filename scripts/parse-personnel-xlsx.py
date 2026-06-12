from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip()) if value is not None else ""


def is_index(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float) and value.is_integer():
        return True
    return bool(re.fullmatch(r"\d+", text(value)))


def department_from_row(cells: list[str], current: str) -> str:
    line = text(" ".join(cell for cell in cells if cell))
    if not line:
        return current
    if "เกษตรศาสตร์" in line:
        return "คณะเกษตรศาสตร์และเทคโนโลยี"
    if "เทคโนโลยีการจัดการ" in line:
        return "คณะเทคโนโลยีการจัดการ"
    if "สำนักงานวิทยาเขต" in line:
        return "สำนักงานวิทยาเขตสุรินทร์"
    return current


def parse_name(full_name: str) -> dict[str, str | None]:
    normalized = text(full_name)
    prefixes = [
        "ว่าที่ ร.ต.",
        "ว่าที่ร.ต.",
        "ส.อ.",
        "นางสาว",
        "นาย",
        "นาง",
        "MRS.",
        "Mrs.",
        "MISS",
        "Miss",
        "MR.",
        "Mr.",
    ]
    title = None
    name = normalized
    for prefix in prefixes:
        if normalized.startswith(prefix):
            title = prefix
            name = text(normalized[len(prefix) :])
            break

    parts = name.split(" ")
    if len(parts) <= 1:
        first_name = name
        last_name = "-"
    else:
        first_name = " ".join(parts[:-1])
        last_name = parts[-1]

    return {
        "title_th": title,
        "first_name_th": first_name,
        "last_name_th": last_name,
    }


def make_employee_code(sheet_index: int, row_index: int, position_no: str) -> str:
    safe_no = re.sub(r"[^0-9A-Za-zก-๙_.()-]+", "-", position_no)[:24]
    suffix = safe_no or f"ROW{row_index}"
    return f"HR66-S{sheet_index:02d}-R{row_index:04d}-{suffix}"[:50]


def make_email(employee_code: str) -> str:
    safe = re.sub(r"[^0-9A-Za-z]+", ".", employee_code.lower()).strip(".")
    return f"personnel.{safe}@example.local"


def parse_workbook(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    records: list[dict[str, object]] = []

    for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
        if worksheet.title.startswith("รายงาน"):
            continue

        current_department = "สำนักงานวิทยาเขตสุรินทร์"
        category = text(worksheet.cell(1, 1).value) or worksheet.title

        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            cells = [text(value) for value in row]
            current_department = department_from_row(cells[:8], current_department)

            if not cells or not is_index(row[0]):
                continue

            full_name = ""
            position_no = ""
            position = ""
            education = ""

            if len(cells) >= 5 and cells[2] and cells[3] and "ชื่อ" not in cells[2]:
                full_name = text(cells[2])
                position_no = text(cells[1])
                position = text(cells[3])
                education = text(cells[4]) if len(cells) > 4 else ""
            elif len(cells) >= 8 and cells[2] in {"นาย", "นาง", "นางสาว"} and cells[3] and cells[4]:
                full_name = text(f"{cells[2]}{cells[3]} {cells[4]}")
                position_no = text(cells[1])
                position = text(cells[6] or cells[5])
                education = text(cells[7]) if len(cells) > 7 else ""

            if not full_name or not position:
                continue

            employee_code = make_employee_code(sheet_index, row_index, position_no)
            name_parts = parse_name(full_name)
            records.append(
                {
                    "employee_code": employee_code,
                    "email": make_email(employee_code),
                    "full_name": full_name,
                    **name_parts,
                    "position_th": position,
                    "department_name_th": current_department,
                    "source_sheet": worksheet.title,
                    "source_row": row_index,
                    "source_position_no": position_no,
                    "personnel_category": category,
                    "education": education,
                }
            )

    return records


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: parse-personnel-xlsx.py <workbook.xlsx>")
    records = parse_workbook(Path(sys.argv[1]))
    json.dump(records, sys.stdout, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
